"""
Export SMS manual test cases to Excel.
Run: python export_test_cases_to_excel.py
Output: SOLUTO/RelatedSystem/test-cases/SMS_Manual_Test_Cases.xlsx
"""
import csv
import os

# Try openpyxl for .xlsx; fallback to CSV
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "RelatedSystem", "test-cases")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Test cases data: (tc_id, title, scenario, priority, steps_list, expected_overall, test_data_dict)
# steps_list = [(step_num, action, expected_result), ...]
def get_test_cases_data():
    return [
        {
            "tc_id": "SMS-CSV-WELCOME-001",
            "title": "Verify Welcome and all three messages for CSV-enrolled MDN (AU), then re-enrollment with same MDN",
            "scenario": "C00553",
            "priority": "High",
            "expected_overall": "First enrollment: SUCCESS, Welcome delivered, all three messages sent. Re-enrollment (same MDN): CONTRACT_TRANSITIONED, Welcome delivered again, all three messages sent again.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "true", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a new SMS contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=true, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN.", "Contract data ready."),
                (2, "Create the CSV file containing this contract/MDN in the format required by the system for enrollment.", "CSV file created."),
                (3, "Upload the CSV file to the configured S3 bucket path used for enrollment.", "Upload successful."),
                (4, "Wait for the enrollment job to run, then wait up to ten minutes for processing to complete.", "Enrollment completes."),
                (5, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC; Locate the row for the test MDN.", "New enrollment record visible for the MDN."),
                (6, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<your_mdn>' ORDER BY created_datetime DESC LIMIT 10; Replace <your_mdn> with the MDN used in Step 1. Note the contract_id.", "Contract ID visible; contract_id noted."),
                (7, "Wait 20 seconds (20000 ms).", "Time for downstream processing."),
                (8, "In Titan DB, verify status SUCCESS for the contract. Run: SELECT contract_id, status, created_datetime FROM <titan_contract_status_table> WHERE contract_id = '<contract_id>' ORDER BY created_datetime DESC;", "Status = SUCCESS."),
                (9, "In ISS DB, verify the Welcome message for the new user. Run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100;", "Welcome message present for the contract."),
                (10, "Wait 90 seconds (90000 ms).", "Time for message delivery."),
                (11, "In Titan DB, get the message history ID for this contract. Run: SELECT * FROM titan_views.sms_system_message_realtime_view WHERE contract_id = '<contract_id>' ORDER BY created_datetime DESC LIMIT 10;", "Message history ID obtained."),
                (12, "In Titan DB, verify all three message requests (Welcome, Remind1, Remind2) for this contract. Run: SELECT * FROM carrier_integration.message_target WHERE contract_id = '<contract_id>' ORDER BY created_datetime DESC;", "All three message requests present/sent."),
                (13, "Using the applicable admin or API process, delete the MDN that was created in Step 1.", "MDN deleted."),
                (14, "Using the applicable process, re-add the same MDN (same phone number) for SMS.", "Same MDN re-added."),
                (15, "Create a new CSV file containing the same MDN and the required contract fields. Upload this CSV to the S3 bucket.", "CSV created and uploaded."),
                (16, "Wait for the enrollment job to run, then wait up to ten minutes for processing.", "Re-enrollment completes."),
                (17, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC;", "Re-enrollment record visible."),
                (18, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<same_mdn>' ORDER BY created_datetime DESC LIMIT 10; Note the contract_id.", "Contract ID for re-enrolled MDN visible."),
                (19, "Wait 5 seconds (5000 ms).", "Time for status update."),
                (20, "In Titan DB, verify status CONTRACT_TRANSITIONED for the re-enrolled contract. Same query pattern as Step 8 with the re-enrolled contract_id.", "Status = CONTRACT_TRANSITIONED."),
                (21, "In ISS DB, verify the Welcome message is delivered again for the re-enrolled user. Same query as Step 9.", "Welcome message present for re-enrolled contract."),
                (22, "Wait 5 seconds (5000 ms).", "Time for messages."),
                (23, "In Titan DB, get the message history ID for the re-enrolled contract. Same query pattern as Step 11.", "Message history ID obtained."),
                (24, "In Titan DB, verify all three message requests for the re-enrolled contract. Same query pattern as Step 12.", "All three message requests present/sent."),
            ]
        },
        {
            "tc_id": "SMS-CSV-REMIND1-002",
            "title": "Verify Reminder One message is delivered to the newly subscribed user after CSV enrollment (AU).",
            "scenario": "C00554",
            "priority": "High",
            "expected_overall": "Reminder One message is delivered to the newly subscribed user (ISS DB).",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "true", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a new SMS contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=true, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN.", "Contract data ready."),
                (2, "Create the CSV file containing this contract/MDN in the format required by the system for enrollment.", "CSV file created."),
                (3, "Upload the CSV file to the configured S3 bucket path used for enrollment.", "Upload successful."),
                (4, "Wait for the enrollment job to run, then wait up to ten minutes for processing to complete.", "Enrollment completes."),
                (5, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC; Locate the row for the test MDN.", "New enrollment record visible."),
                (6, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<mdn>' ORDER BY created_datetime DESC LIMIT 10; Replace <mdn> with the MDN from Step 1. Note the contract_id.", "Contract ID visible."),
                (7, "Wait 90 seconds (90000 ms) to allow Reminder One to be sent.", "Wait complete."),
                (8, "In ISS DB, verify the Reminder One message for the newly subscribed user. Run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Filter or identify the row for this contract/user and message type Reminder One.", "Reminder One message delivered and visible in ISS DB."),
            ]
        },
        {
            "tc_id": "SMS-CSV-REMIND2-003",
            "title": "Verify Reminder Two message is delivered to the newly subscribed user after CSV enrollment (AU).",
            "scenario": "C00641",
            "priority": "High",
            "expected_overall": "Reminder Two message is delivered to the newly subscribed user (ISS DB).",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "true", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a new SMS contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=true, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN.", "Contract data ready."),
                (2, "Create the CSV file containing this contract/MDN in the format required by the system for enrollment.", "CSV file created."),
                (3, "Upload the CSV file to the configured S3 bucket path used for enrollment.", "Upload successful."),
                (4, "Wait for the enrollment job to run, then wait up to ten minutes for processing to complete.", "Enrollment completes."),
                (5, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC; Locate the row for the test MDN.", "New enrollment record visible."),
                (6, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<mdn>' ORDER BY created_datetime DESC LIMIT 10; Replace <mdn> with the MDN from Step 1. Note the contract_id.", "Contract ID visible."),
                (7, "Wait 90 seconds (90000 ms) to allow Reminder Two to be sent.", "Wait complete."),
                (8, "In ISS DB, verify the Reminder Two message for the newly subscribed user. Run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Filter or identify the row for this contract/user and message type Reminder Two.", "Reminder Two message delivered and visible in ISS DB."),
            ]
        },
        {
            "tc_id": "SMS-TITAN-OPENSEARCH-004",
            "title": "Enroll an MDN via CSV, then verify Welcome/Remind One/Remind Two SMS delivered IDs match Titan DB and contract ID appears in Titan OpenSearch Dashboard.",
            "scenario": "C00651",
            "priority": "High",
            "expected_overall": "Welcome/Remind One/Remind Two SMS delivered IDs match Titan DB SMS logs, and the contract ID is visible in Titan OpenSearch Dashboard.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "true", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a new SMS contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=true, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique values.", "Contract data ready."),
                (2, "Create the CSV file containing this contract/MDN and upload it to the S3 bucket.", "CSV created and uploaded."),
                (3, "Wait for the enrollment job to run, then wait up to ten minutes for processing to complete.", "Enrollment completes."),
                (4, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC; Locate the row for the test MDN.", "Enrollment record visible."),
                (5, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<mdn>' ORDER BY created_datetime DESC LIMIT 10; Replace <mdn> with the MDN from Step 1. Note the contract_id.", "Contract ID obtained."),
                (6, "In Titan DB, get the SMS log data for this contract (delivered IDs for Welcome, Remind One, Remind Two). Run: SELECT * FROM titan_views.sms_system_message_realtime_view WHERE contract_id = '<contract_id>' ORDER BY created_datetime DESC LIMIT 10; Note the SMS delivered IDs.", "SMS delivered IDs obtained from Titan DB."),
                (7, "Open the web app and navigate to the Titan Open Search Dashboard application.", "On Titan OpenSearch Dashboard."),
                (8, "Log in via SSO.", "Logged in."),
                (9, "Compare the SMS delivered IDs for Welcome, Remind One, and Remind Two from Step 6 with the logs displayed in Titan OpenSearch. Verify they match.", "Welcome/Remind1/Remind2 delivered IDs match Titan DB SMS logs."),
                (10, "In the Titan OpenSearch Dashboard, enter a query using the Contract Id (from Step 5) in the search text field and execute the search.", "Query executed."),
                (11, "Verify that the contract id appears in the logs in the Titan OpenSearch Dashboard.", "Contract ID is present in Titan OpenSearch logs."),
            ]
        },
        {
            "tc_id": "SMS-HORIZON-OPENSEARCH-005",
            "title": "Enroll an MDN via CSV, then verify Welcome/Remind One/Remind Two SMS delivered IDs match Titan DB and Enrollment/Welcome SMS logs visible in Horizon OpenSearch Dashboard.",
            "scenario": "C00653",
            "priority": "High",
            "expected_overall": "Welcome/Remind One/Remind Two delivered IDs match Titan DB, and Enrollment/Welcome SMS logs are visible in Horizon OpenSearch Dashboard.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "true", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a new SMS contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=true, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique values.", "Contract data ready."),
                (2, "Create the CSV file containing this contract/MDN and upload it to the S3 bucket.", "CSV created and uploaded."),
                (3, "Wait for the enrollment job to run, then wait up to ten minutes for processing to complete.", "Enrollment completes."),
                (4, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC; Locate the row for the test MDN.", "Enrollment record visible."),
                (5, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<mdn>' ORDER BY created_datetime DESC LIMIT 10; Replace <mdn> with the MDN from Step 1. Note the contract_id and subscriber_id.", "Contract ID and subscriber ID obtained."),
                (6, "In Titan DB, get the SMS delivered IDs for Welcome, Remind One, and Remind Two. Run: SELECT * FROM titan_views.sms_system_message_realtime_view WHERE contract_id = '<contract_id>' ORDER BY created_datetime DESC LIMIT 10;", "Delivered IDs obtained from Titan DB."),
                (7, "Open the web app and navigate to the Horizon Open Search Dashboard application.", "On Horizon OpenSearch Dashboard."),
                (8, "Log in via SSO.", "Logged in."),
                (9, "Verify that the SMS delivered IDs for Welcome, Remind One, and Remind Two match the Titan DB SMS logs (from Step 6).", "IDs match Titan DB."),
                (10, "In the Horizon OpenSearch Dashboard, enter a query using the Subscriber Id (from Step 5) in the search text field and execute the search.", "Query executed."),
                (11, "Verify that Enrollment / Welcome SMS logs are visible in the Horizon OpenSearch Dashboard.", "Enrollment/Welcome SMS logs visible in Horizon dashboard."),
            ]
        },
        {
            "tc_id": "SMS-CALL-RESERVE-009",
            "title": "Verify Call reservation confirmation SMS is sent after reserving a time slot (unprovisioned MDN, AU).",
            "scenario": "C00557",
            "priority": "High",
            "expected_overall": "Call reservation confirmation SMS transmission is completed.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "false", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a SMS new contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=false, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN. This represents an unprovisioned MDN for call reservation.", "Contract/MDN created."),
                (2, "Using the system's call reservation flow (UI or API), reserve a time slot with this new MDN for call reservation.", "Time slot reserved."),
                (3, "Wait 10 seconds (10000 ms).", "Time for SMS to be sent."),
                (4, "Verify that Call reservation SMS transmission is completed. In ISS DB run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Identify the Call reservation confirmation SMS for this MDN/contract.", "Call reservation confirmation SMS sent and verifiable."),
            ]
        },
        {
            "tc_id": "SMS-CALL-REMIND-010",
            "title": "Verify Call reservation reminder SMS is sent after reserving a time slot (unprovisioned MDN, AU).",
            "scenario": "C00558",
            "priority": "High",
            "expected_overall": "Call reservation reminder SMS transmission is completed.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "false", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a SMS new contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=false, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN.", "Contract/MDN created."),
                (2, "Using the system's call reservation flow, reserve a time slot with this new MDN for call reservation.", "Time slot reserved."),
                (3, "Wait 35 seconds (35000 ms) to allow the reminder SMS to be sent.", "Wait complete."),
                (4, "Verify Call reminder SMS transmission is completed. In ISS DB run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Identify the Call reservation reminder SMS for this MDN/contract.", "Call reminder SMS sent and verifiable."),
            ]
        },
        {
            "tc_id": "SMS-CALL-CANCEL-011",
            "title": "Verify Call cancellation SMS is sent after cancelling a call reservation (unprovisioned MDN, AU).",
            "scenario": "C00559",
            "priority": "High",
            "expected_overall": "Call cancellation SMS transmission is completed.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "false", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a SMS new contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=false, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN.", "Contract/MDN created."),
                (2, "Using the system's call reservation flow, reserve a time slot with this new MDN for call reservation.", "Time slot reserved."),
                (3, "Wait 45 seconds (45000 ms).", "Time before cancel."),
                (4, "Using the system's cancel flow (UI or API), cancel the call reservation for this MDN.", "Reservation cancelled."),
                (5, "Wait 45 seconds (45000 ms) for the cancellation SMS to be sent.", "Wait complete."),
                (6, "Verify Call cancellation SMS transmission is completed. In ISS DB run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Identify the Call cancellation SMS for this MDN/contract.", "Call cancellation SMS sent and verifiable."),
            ]
        },
        {
            "tc_id": "SMS-CALL-NOT-SENT-PROVISIONED-012",
            "title": "Verify Call reservation SMS is not sent when the user is provisioned as auNonSubscriber (AU).",
            "scenario": "C00664",
            "priority": "High",
            "expected_overall": "Call reservation SMS is not sent to the provisioned user.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "false", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Create a SMS new contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=false, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique subscriberId and MDN.", "Contract/MDN created."),
                (2, "Using the project's provisioning process, provision the user as new auNonSubscriber for this MDN.", "User provisioned as auNonSubscriber."),
                (3, "Using the system's call reservation flow, reserve a time slot with this new MDN for call reservation.", "Time slot reserved."),
                (4, "Wait 35 seconds (35000 ms).", "Time for SMS window."),
                (5, "Verify that Call reservation SMS is not sent to this provisioned user. In ISS DB run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Confirm there is no Call reservation SMS for this MDN/contract.", "Call reservation SMS is not sent to the provisioned user."),
            ]
        },
        {
            "tc_id": "SMS-MANUAL-DELIVERY-015",
            "title": "With Manual delivery ON (isNormalOperationMode=false), enroll via CSV, verify SMS in TO_BE_PROCESSED, approve re-delivery, then verify DELIVERED in Titan and Welcome in ISS.",
            "scenario": "C10000",
            "priority": "High",
            "expected_overall": "With Manual delivery ON: SMS in TO_BE_PROCESSED; after re-delivery approval, all three show DELIVERED in Titan and Welcome is delivered in ISS. No duplicate logs; isNormalOperationMode set back to true.",
            "test_data": {"subscriberId": "dynamic_Subscriber", "mdn": "dynamic_MDN", "serviceEnabled": "true", "kfsServiceEnabled": "true", "appleSubscriptionEnabled": "true", "birthday": "19700101", "gender": "female", "brand": "AU"},
            "steps": [
                (1, "Call the isNormalOperationMode endpoint and set the value to \"false\" in the JSON body.", "Request accepted; Manual delivery ON."),
                (2, "Verify that manual delivery mode of SMS is ON when isNormalOperationMode is \"false\" (per project verification step).", "Manual delivery mode = ON."),
                (3, "Create a SMS new contract with: subscriberId (e.g. dynamic_Subscriber), MDN (e.g. dynamic_MDN), serviceEnabled=true, kfsServiceEnabled=true, appleSubscriptionEnabled=true, birthday=19700101, gender=female, brand=AU. Use unique values.", "Contract data ready."),
                (4, "Create the CSV file containing this contract/MDN and upload it to the S3 bucket.", "CSV created and uploaded."),
                (5, "Wait for the enrollment job to run, then wait up to ten minutes for processing to complete.", "Enrollment completes."),
                (6, "In the carrier enrollment database, run: SELECT * FROM carrier_integration.message_target mt ORDER BY created_datetime DESC; Locate the row for the test MDN.", "Enrollment record visible."),
                (7, "In the customer contracts database, run: SELECT * FROM customer_contracts WHERE mdn = '<mdn>' ORDER BY created_datetime DESC LIMIT 10; Replace <mdn> with the MDN from Step 3. Note the contract_id.", "Contract ID visible."),
                (8, "Wait 40 seconds (40000 ms).", "Processing time."),
                (9, "In Titan DB, for case \"1\", verify status SUCCESS for the enrolled contract. Run: SELECT contract_id, status, created_datetime FROM <titan_contract_status_table> WHERE contract_id = '<contract_id>' ORDER BY created_datetime DESC;", "Status = SUCCESS."),
                (10, "In ISS DB, verify that Welcome, Remind One, and Remind Two are not delivered yet (manual mode). Run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100;", "No Welcome/Remind1/Remind2 in ISS."),
                (11, "In Titan DB, verify that all three SMS delivered logs have status TO_BE_PROCESSED. Use the project's Titan SMS delivery status table or view.", "All three SMS in TO_BE_PROCESSED."),
                (12, "Check for duplicate logs in Titan DB (run the project's duplicate-check query for Titan).", "No duplicate logs."),
                (13, "Check for duplicate logs using a GET request to the issMock endpoint (use the project's URL and parameters).", "No duplicates in issMock."),
                (14, "Send a POST request to the endpoint that approves re-delivery so that all three SMS delivered status will be updated and reflected in ISS DB.", "Re-delivery approved."),
                (15, "Wait 40 seconds (40000 ms).", "Processing time."),
                (16, "In Titan DB, verify that all three SMS delivered logs have status DELIVERED. Use the same Titan SMS delivery status table or view as in Step 11.", "All three SMS = DELIVERED."),
                (17, "Wait 10 seconds (10000 ms).", "Processing time."),
                (18, "In ISS DB, verify the Welcome message is delivered to the newly subscribed user. Run: SELECT * FROM sms_system.message ORDER BY created_datetime DESC LIMIT 100; Identify the Welcome message for the contract from Step 7.", "Welcome message visible in ISS DB."),
                (19, "Call the isNormalOperationMode endpoint and set the value to \"true\" in the JSON body to restore normal operation.", "Normal operation restored."),
            ]
        },
    ]


def write_excel(out_path):
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.append(["Test Case ID", "Title", "Scenario", "Priority", "Expected Result (Overall)"])
    for r in ws1[1]:
        r.fill = header_fill
        r.font = header_font
    for tc in get_test_cases_data():
        ws1.append([tc["tc_id"], tc["title"], tc["scenario"], tc["priority"], tc["expected_overall"]])
    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 25
    # Sheet 2: Steps (all steps in one sheet)
    ws2 = wb.create_sheet("Steps")
    ws2.append(["Test Case ID", "Step", "Action", "Expected Result"])
    for r in ws2[1]:
        r.fill = header_fill
        r.font = header_font
    for tc in get_test_cases_data():
        for step_num, action, expected in tc["steps"]:
            ws2.append([tc["tc_id"], step_num, action, expected])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 80
    ws2.column_dimensions["D"].width = 50
    # Sheet 3: Test Data
    ws3 = wb.create_sheet("Test Data")
    keys = ["Test Case ID", "subscriberId", "mdn", "serviceEnabled", "kfsServiceEnabled", "appleSubscriptionEnabled", "birthday", "gender", "brand"]
    ws3.append(keys)
    for r in ws3[1]:
        r.fill = header_fill
        r.font = header_font
    for tc in get_test_cases_data():
        row = [tc["tc_id"]] + [tc["test_data"].get(k, "") for k in keys[1:]]
        ws3.append(row)
    for col in range(1, len(keys) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 22
    wb.save(out_path)
    print(f"Excel file created: {out_path}")


def write_csv(out_path):
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Test Case ID", "Title", "Scenario", "Priority", "Step", "Action", "Expected Result", "Expected Result (Overall)"])
        for tc in get_test_cases_data():
            overall = tc["expected_overall"]
            for i, (step_num, action, expected) in enumerate(tc["steps"]):
                title = tc["title"] if i == 0 else ""
                scenario = tc["scenario"] if i == 0 else ""
                priority = tc["priority"] if i == 0 else ""
                tc_id = tc["tc_id"] if i == 0 else ""
                w.writerow([tc_id, title, scenario, priority, step_num, action, expected, overall])
    print(f"CSV file created (open in Excel): {out_path}")


if __name__ == "__main__":
    xlsx_path = os.path.join(OUTPUT_DIR, "SMS_Manual_Test_Cases.xlsx")
    csv_path = os.path.join(OUTPUT_DIR, "SMS_Manual_Test_Cases.csv")
    if HAS_OPENPYXL:
        write_excel(xlsx_path)
    else:
        print("openpyxl not installed. Install with: pip install openpyxl")
        print("Writing CSV instead (you can open it in Excel and Save As .xlsx).")
        write_csv(csv_path)
